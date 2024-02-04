# print hello world
print("Hello World")

# print todays date
import datetime
print(datetime.date.today())

# read in excel file and print all cells with data in them
import openpyxl as xl   
wb = xl.load_workbook('test.xlsx')
sheet = wb['Sheet1']
for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, col)
        print(cell.value)
ss+


